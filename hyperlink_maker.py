import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import openpyxl
from openpyxl.styles import Font
import os
import sys
import platform


if platform.system() == "Windows":
    try:
        import windnd
        WINDND_AVAILABLE = True
    except ImportError:
        WINDND_AVAILABLE = False
else:
    WINDND_AVAILABLE = False



class HyperlinkMakerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Hyperlink Maker")
        self.root.minsize(700, 650)
        self.root.geometry("700x650")
        self.root.configure(bg="#e3f2fd")
        
        icon_path = os.path.join(os.path.dirname(__file__), "Hyperlink_Maker_Icon_256.ico")
        if os.path.exists(icon_path):
            self.root.iconbitmap(icon_path)

        self.excel_file = None
        self.workbook = None
        self.selected_columns = []
        
        self.setup_styles()
        self.create_widgets()
        
    def setup_styles(self):
        style = ttk.Style()
        style.theme_use("clam")
        
        style.configure("TFrame", background="#e3f2fd")
        
        style.configure("Title.TLabel", 
                       background="#e3f2fd", 
                       foreground="#1565c0", 
                       font=("Segoe UI", 22, "bold"))
        
        style.configure("TLabel", 
                       background="#e3f2fd", 
                       foreground="#0d47a1", 
                       font=("Segoe UI", 11))
        
        style.configure("Header.TLabel",
                       background="#e3f2fd",
                       foreground="#1565c0",
                       font=("Segoe UI", 12, "bold"))
        
        style.configure("TButton",
                       background="#1976d2",
                       foreground="#ffffff",
                       font=("Segoe UI", 11, "bold"),
                       padding=12)
        style.map("TButton",
                 background=[("active", "#1565c0")])
        
        style.configure("Status.TLabel",
                       background="#e3f2fd",
                       foreground="#388e3c",
                       font=("Segoe UI", 10))
        
    def create_widgets(self):
        main_frame = ttk.Frame(self.root, style="TFrame")
        main_frame.pack(fill=tk.BOTH, expand=True, padx=30, pady=15)
        
        title = ttk.Label(main_frame, text="Hyperlink Maker", style="Title.TLabel")
        title.pack(pady=(0, 2))
        
        subtitle = ttk.Label(main_frame, text="Convert Excel columns to clickable hyperlinks")
        subtitle.pack(pady=(0, 10))
        
        self.drop_frame = tk.Frame(main_frame, bg="#bbdefb", height=80, width=640, 
                                   highlightthickness=3, highlightbackground="#1976d2",
                                   relief=tk.SOLID)
        self.drop_frame.pack_propagate(False)
        self.drop_frame.pack(pady=(0, 12))
        
        self.drop_inner = tk.Frame(self.drop_frame, bg="#bbdefb")
        self.drop_inner.place(relx=0.5, rely=0.5, anchor=tk.CENTER)
        
        self.drop_label = tk.Label(self.drop_inner, 
                             text="Drag & Drop Excel File Here",
                             bg="#bbdefb", 
                             fg="#0d47a1",
                             font=("Segoe UI", 13, "bold"))
        self.drop_label.pack()
        
        self.drop_label2 = tk.Label(self.drop_inner, 
                              text="or click the button below",
                              bg="#bbdefb", 
                              fg="#1976d2",
                              font=("Segoe UI", 10))
        self.drop_label2.pack()
        
        if WINDND_AVAILABLE:
            windnd.hook_dropfiles(self.drop_frame, func=self.handle_drop)
            windnd.hook_dropfiles(self.drop_inner, func=self.handle_drop)
            windnd.hook_dropfiles(self.drop_label, func=self.handle_drop)
            windnd.hook_dropfiles(self.drop_label2, func=self.handle_drop)
        
        select_btn = ttk.Button(main_frame, text="Select Excel File", command=self.select_file)
        select_btn.pack(pady=(0, 12))
        
        self.file_frame = tk.Frame(main_frame, bg="#e3f2fd")
        self.file_frame.pack(pady=(0, 8), fill=tk.X)
        
        self.file_label = ttk.Label(self.file_frame, text="No file selected", 
                                    style="TLabel", font=("Segoe UI", 10, "italic"))
        self.file_label.pack()
        
        self.sheet_frame = ttk.Frame(main_frame, style="TFrame")
        self.sheet_frame.pack(pady=(0, 8), fill=tk.X)
        
        sheet_label = ttk.Label(self.sheet_frame, text="Select sheet:", 
                                  style="Header.TLabel")
        sheet_label.pack(anchor=tk.W, pady=(0, 5))
        
        self.sheet_combo = ttk.Combobox(self.sheet_frame, state="readonly", 
                                         font=("Segoe UI", 11), width=40)
        self.sheet_combo.pack(anchor=tk.W, fill=tk.X)
        self.sheet_combo.bind("<<ComboboxSelected>>", self.on_sheet_changed)
        
        self.columns_frame = ttk.Frame(main_frame, style="TFrame")
        self.columns_frame.pack(pady=(0, 8), fill=tk.BOTH, expand=True)
        
        columns_label = ttk.Label(self.columns_frame, text="Select columns to hyperlink:", 
                                  style="Header.TLabel")
        columns_label.pack(anchor=tk.W, pady=(0, 6))
        
        self.columns_container = tk.Frame(self.columns_frame, bg="#ffffff", 
                                           highlightthickness=2, highlightbackground="#1976d2",
                                           height=100)
        self.columns_container.pack(fill=tk.BOTH, expand=True)
        self.columns_container.pack_propagate(False)
        
        self.columns_scrollbar = ttk.Scrollbar(self.columns_container, orient=tk.VERTICAL)
        self.columns_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.columns_listbox = tk.Listbox(self.columns_container, 
                                          bg="#ffffff",
                                          fg="#0d47a1",
                                          font=("Segoe UI", 11),
                                          selectmode=tk.MULTIPLE,
                                          yscrollcommand=self.columns_scrollbar.set,
                                          highlightthickness=0,
                                          borderwidth=0)
        self.columns_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=2, pady=2)
        self.columns_scrollbar.config(command=self.columns_listbox.yview)
        
        self.select_all_frame = tk.Frame(self.columns_frame, bg="#e3f2fd")
        self.select_all_frame.pack(fill=tk.X, pady=(5, 0))
        
        self.select_all_btn = tk.Button(self.select_all_frame, text="Select All",
                                        bg="#42a5f5", fg="white", font=("Segoe UI", 9),
                                        relief=tk.FLAT, padx=10, pady=3,
                                        command=self.select_all_columns)
        self.select_all_btn.pack(side=tk.LEFT, padx=2)
        
        self.clear_all_btn = tk.Button(self.select_all_frame, text="Clear All",
                                       bg="#90caf9", fg="#0d47a1", font=("Segoe UI", 9),
                                       relief=tk.FLAT, padx=10, pady=3,
                                       command=self.clear_all_columns)
        self.clear_all_btn.pack(side=tk.LEFT, padx=2)
        
        self.progress_frame = tk.Frame(main_frame, bg="#e3f2fd")
        self.progress_frame.pack(fill=tk.X, pady=(0, 5))
        
        self.progress = ttk.Progressbar(self.progress_frame, mode="determinate", 
                                        length=300)
        self.progress.pack(pady=(0, 3))
        
        self.status_label = ttk.Label(self.progress_frame, text="", style="Status.TLabel")
        self.status_label.pack()
        
        button_frame = ttk.Frame(main_frame, style="TFrame")
        button_frame.pack(pady=(5, 8), fill=tk.X)
        
        buttons_inner = tk.Frame(button_frame, bg="#e3f2fd")
        buttons_inner.pack()
        
        self.process_btn = tk.Button(buttons_inner, text="Create Hyperlinks & Replace File", 
                                     bg="#f44336", fg="white", font=("Segoe UI", 10, "bold"),
                                     relief=tk.FLAT, padx=15, pady=8,
                                     command=self.process_file, state=tk.DISABLED,
                                     cursor="arrow")
        self.process_btn.pack(side=tk.LEFT, padx=5, pady=5)
        
        self.save_new_btn = tk.Button(buttons_inner, text="Save As New File",
                                       bg="#4caf50", fg="white", font=("Segoe UI", 10),
                                       relief=tk.FLAT, padx=15, pady=8,
                                       command=self.process_file_new, state=tk.DISABLED,
                                       cursor="arrow")
        self.save_new_btn.pack(side=tk.LEFT, padx=5, pady=5)
    
    def _normalize_dropped_path(self, file_path):
        if file_path is None:
            return None
        if isinstance(file_path, bytes):
            try:
                file_path = file_path.decode("utf-8")
            except UnicodeDecodeError:
                try:
                    file_path = file_path.decode("mbcs")
                except Exception:
                    file_path = file_path.decode(errors="replace")

        file_path = str(file_path).strip()
        if len(file_path) >= 2 and file_path[0] == "{" and file_path[-1] == "}":
            file_path = file_path[1:-1]
        if len(file_path) >= 2 and file_path[0] == '"' and file_path[-1] == '"':
            file_path = file_path[1:-1]
        return file_path

    def handle_drop(self, files):
        if isinstance(files, (list, tuple)):
            if not files:
                return
            file_path = files[0]
        else:
            file_path = files

        file_path = self._normalize_dropped_path(file_path)
        if not file_path:
            return

        if file_path.lower().endswith((".xlsx", ".xls")):
            self.load_file(file_path)
        else:
            messagebox.showerror("Error", "Please drop an Excel file (.xlsx or .xls)")
    
    def select_file(self):
        file_path = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if file_path:
            self.load_file(file_path)
    
    def load_file(self, file_path):
        try:
            self.excel_file = file_path
            self.workbook = openpyxl.load_workbook(file_path)
            
            basename = os.path.basename(file_path)
            self.file_label.config(text=f"Loaded: {basename}")
            
            sheet_names = self.workbook.sheetnames
            self.sheet_combo['values'] = sheet_names
            if sheet_names:
                self.sheet_combo.current(0)
                self.root.after(50, self.show_columns_from_first_sheet)
            
            self.process_btn.config(state=tk.NORMAL, cursor="hand2")
            self.save_new_btn.config(state=tk.NORMAL, cursor="hand2")
            self.status_label.config(text="File loaded successfully!", foreground="#388e3c")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load file: {str(e)}")
    
    def show_columns_from_first_sheet(self):
        sheet_name = self.sheet_combo.get()
        if sheet_name and self.workbook:
            sheet = self.workbook[sheet_name]
            self.load_columns_for_sheet(sheet)
    
    def on_sheet_changed(self, event):
        sheet_name = self.sheet_combo.get()
        if sheet_name and self.workbook:
            sheet = self.workbook[sheet_name]
            self.load_columns_for_sheet(sheet)
    
    def load_columns_for_sheet(self, sheet):
        self.columns_listbox.delete(0, tk.END)
        
        headers = []
        
        for cell in sheet[1]:
            if cell.value:
                try:
                    col_letter = cell.column_letter
                    headers.append((col_letter, str(cell.value)))
                except:
                    pass
        
        self.column_info = {}
        for col_letter, header in headers:
            display_text = f"{header}  [{col_letter}]"
            self.columns_listbox.insert(tk.END, display_text)
            self.column_info[display_text] = col_letter
    
    def select_all_columns(self):
        self.columns_listbox.select_set(0, tk.END)
    
    def clear_all_columns(self):
        self.columns_listbox.select_clear(0, tk.END)
    
    def get_selected_columns(self):
        selected_indices = self.columns_listbox.curselection()
        selected_cols = []
        for idx in selected_indices:
            display_text = self.columns_listbox.get(idx)
            col_letter = self.column_info.get(display_text)
            if col_letter:
                selected_cols.append(col_letter)
        return selected_cols
    
    def process_file(self):
        selected_cols = self.get_selected_columns()
        if not selected_cols:
            messagebox.showwarning("Warning", "Please select at least one column")
            return
        
        if not messagebox.askyesno("Confirm", "This will modify the original file. Continue?"):
            return
        
        save_path = self.excel_file
        self.process_hyperlinks(selected_cols, save_path)
    
    def process_file_new(self):
        selected_cols = self.get_selected_columns()
        if not selected_cols:
            messagebox.showwarning("Warning", "Please select at least one column")
            return
        
        save_path = filedialog.asksaveasfilename(
            title="Save As",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if not save_path:
            return
        
        self.process_hyperlinks(selected_cols, save_path)
    
    def process_hyperlinks(self, columns, save_path):
        sheet_name = self.sheet_combo.get()
        sheet = self.workbook[sheet_name]
        max_row = sheet.max_row
        
        self.progress["maximum"] = max_row
        self.progress["value"] = 0
        
        hyperlink_font = Font(color="0563C1", underline="single")
        
        for row_idx in range(2, max_row + 1):
            for col_letter in columns:
                cell = sheet[f"{col_letter}{row_idx}"]
                if cell.value:
                    cell_value = str(cell.value)
                    if cell_value.startswith(("http://", "https://", "www.")):
                        url = cell_value if cell_value.startswith(("http://", "https://")) else f"https://{cell_value}"
                    else:
                        url = cell_value
                    
                    cell.hyperlink = url
                    cell.font = hyperlink_font
            
            self.progress["value"] = row_idx
            self.status_label.config(text=f"Processing row {row_idx} of {max_row}...")
            self.root.update_idletasks()
        
        try:
            self.workbook.save(save_path)
            self.status_label.config(text=f"Saved: {os.path.basename(save_path)}")
            messagebox.showinfo("Success", "Hyperlinks created successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save file: {str(e)}")


def main():
    root = tk.Tk()
    
    app = HyperlinkMakerApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
